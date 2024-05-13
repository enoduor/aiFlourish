from django.shortcuts import render, redirect
from .models import stuffs
from .addstuffsform import addStuffs
from django.http import HttpResponse
from openpyxl import load_workbook

# Create your views here.
def truncate_text(text):
    full_stops = text.split('.')
    truncated_text = '.'.join(full_stops[:2])
    return truncated_text
def truncate_steps(text):
    full_stops = text.split('.')
    truncated_text = '.'.join(full_stops[:4])
    return truncated_text
def home(request):
    Details = stuffs.objects.all()
    for Detail in Details:
        Detail.description = truncate_text(Detail.description)
        Detail.tutorial_tool = truncate_steps(Detail.tutorial_tool)
        Detail.ad = truncate_text(Detail.ad)
    return render(request,'index.html',{'Details':Details})

def addStuffsView(request):
    form = addStuffs()
    if request.method == "POST":
        form = addStuffs(request.POST)
        if form.is_valid():
            model_instance = stuffs(
                name_of_tool = form.cleaned_data['name_of_tool'],
                description =  form.cleaned_data['description'],
                tutorial_tool =  form.cleaned_data['tutorial_tool'],
                ad =  form.cleaned_data['ad'],
                youtube_link =  form.cleaned_data['youtube_link'],
            )
            model_instance.save()
            return HttpResponse('Tool Added successfully')
        else:
            form = addStuffs()

    return render (request, 'Addstuffsform.html', {'form':form})

def import_from_excel(request): 
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active
#min_row = 2, 
        for row in ws.iter_rows(min_row = 2, values_only=True):
            name, description, sign_up, ad, video, web_link = row
            stuffs.objects.create(name_of_tool = name, description=description,tutorial_tool=sign_up, ad=ad, youtube_link=video, website_link=web_link)

        return render(request, 'import_success.html')#when inside the for loop, it only returns one row.
    
    return render(request, 'import_form.html')
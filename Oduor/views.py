from django.shortcuts import render, redirect, get_object_or_404
from .models import stuffs, category
from .addstuffsform import addStuffs
from django.http import HttpResponse
from openpyxl import load_workbook
from django.db.models import Q

# viewsclear


def truncate_text(text):
    full_stops = text.split('.')
    truncated_text = '.'.join(full_stops[:10])
    return truncated_text


def truncate_steps(text):
    full_stops = text.split('.')
    truncated_text = '.'.join(full_stops[:4])
    return truncated_text


def home(request):
    Details = stuffs.objects.all()

    for Detail in Details:
        Detail.description = truncate_text(Detail.description)
        Detail.tutorial_tool = truncate_steps(Detail.tutorial_tool)
        Detail.ad = Detail.ad


    categories = category.objects.all()
    selected_categories = request.GET.getlist('categories[]', [])
     # count functionality
    t_count = stuffs.objects.count()
    count = t_count + 850

    if selected_categories:
        Details = Details.filter(categories__name__in=selected_categories).distinct()
        c_count = Details.count()
    else:
        c_count = count

    # Search functionality for Stuffs
    search_query = request.GET.get('search')

    if search_query:
        Details = Details.filter(Q(name_of_tool__icontains=search_query) )

   

    context = {'Details': Details, 'categories': categories,'count':count, 'c_count':c_count  }
    return render(request,'index.html',context)#fetch both stuffs and categories

def filtered_tools(request):
    Details = stuffs.objects.all()
    selected_categories = request.GET.getlist('categories[]', [])
    # count functionality
    t_count = stuffs.objects.count()
    count = t_count + 850

    if selected_categories:
        Details = Details.filter(categories__name__in=selected_categories).distinct()
        c_count = Details.count()
    else:
        c_count = count

    search_query = request.GET.get('search')

    if search_query:
        Details = Details.filter(Q(name_of_tool__icontains=search_query))

    context = {'Details': Details, 'count':count, 'c_count':c_count}
    return render(request, 'selected-toools.html', context)


def addStuffsView(request):
    form = addStuffs()


    if request.method == "POST":
        form = addStuffs(request.POST)


        if form.is_valid():
            model_instance = stuffs(
                name_of_tool = form.cleaned_data['name_of_tool'],
                categories= form.cleaned_data['categories'],
                description =  form.cleaned_data['description'],
                tutorial_tool =  form.cleaned_data['sign_up'],
                #ad =  form.cleaned_data['ad'],
                youtube_link =  form.cleaned_data['youtube_link'],
                website_link = form.cleaned_data['website_link'],
            )
            model_instance.save()
            return HttpResponse('Tool Added successfully')


        else:
            form = addStuffs()


    return render (request, 'Addstuffsform.html', {'form':form})


def import_from_excel(request): 


    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active


        #min_row = 2, 


        for row in ws.iter_rows(min_row = 2, values_only=True):
            name, description, sign_up, ad, video, web_link = row
            stuffs.objects.create(name_of_tool = name, description=description,tutorial_tool=sign_up, ad=ad, youtube_link=video, website_link=web_link)


        return render(request, 'import_success.html')#when inside the for loop, it only returns one row.
    

    return render(request, 'import_form.html')


def ToolDetail(request, stuff_id):


    tool_details = stuffs.objects.get(id=stuff_id)
    tool_details.description = truncate_text(tool_details.description)
    tool_details.tutorial_tool = truncate_steps(tool_details.tutorial_tool)
    return render(request, 'tool.html', {'tool_details':tool_details})


def VideoView(request):
    Details = stuffs.objects.all()
    categories = category.objects.all()
    selected_categories = request.GET.getlist('categories[]', [])

    t_count = stuffs.objects.count()
    count = t_count + 850

    # if selected_categories:
    #     Details = Details.filter(categories__name__in=selected_categories).distinct()

    if selected_categories:
        Details = Details.filter(categories__name__in=selected_categories).distinct()
        c_count = Details.count()
    else:
        c_count = count

    
    # Search functionality for Stuffs
    search_query = request.GET.get('search')

    if search_query:
        Details = Details.filter(Q(name_of_tool__icontains=search_query) )

    context = {'Details': Details, 'categories': categories, 'count':count, 'c_count':c_count }
    return render(request,'video.html',context)

def filtered_video(request):
    Details = stuffs.objects.all()
    selected_categories = request.GET.getlist('categories[]', [])
    # count functionality
    t_count = stuffs.objects.count()
    count = t_count + 850

    if selected_categories:
        Details = Details.filter(categories__name__in=selected_categories).distinct()
        c_count = Details.count()
    else:
        c_count = count

    search_query = request.GET.get('search')

    if search_query:
        Details = Details.filter(Q(name_of_tool__icontains=search_query))

    context = {'Details': Details, 'count':count, 'c_count':c_count}
    return render(request, 'filtered_video.html', context)

